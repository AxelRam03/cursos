from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io


def generate_ticket_pdf_bytes(sale) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(8 * cm, 22 * cm),
        topMargin=0.3 * cm,
        bottomMargin=0.3 * cm,
        leftMargin=0.3 * cm,
        rightMargin=0.3 * cm,
    )

    story = []
    center = ParagraphStyle("center", alignment=TA_CENTER, fontSize=9)
    bold_center = ParagraphStyle("bold_center", alignment=TA_CENTER, fontSize=11, fontName="Helvetica-Bold")
    right = ParagraphStyle("right", alignment=TA_RIGHT, fontSize=9)
    bold_right = ParagraphStyle("bold_right", alignment=TA_RIGHT, fontSize=11, fontName="Helvetica-Bold")

    story.append(Paragraph(sale.store.name if sale.store else "AlaShop", bold_center))
    story.append(Paragraph(f"Folio: {sale.folio}", center))
    story.append(Paragraph(sale.created_at.strftime("%d/%m/%Y %H:%M"), center))
    story.append(HRFlowable(width="100%", thickness=1))
    story.append(Spacer(1, 0.2 * cm))

    data = [["Producto", "Cant.", "P.U.", "Total"]]
    for item in sale.items:
        data.append([
            item.product_name[:14],
            str(item.quantity),
            f"${item.unit_price:.2f}",
            f"${item.subtotal:.2f}",
        ])

    table = Table(data, colWidths=[3 * cm, 1 * cm, 1.7 * cm, 1.7 * cm])
    table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.black),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
    ]))
    story.append(table)
    story.append(HRFlowable(width="100%", thickness=1))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph(f"Subtotal: ${sale.subtotal:.2f}", right))
    if sale.discount > 0:
        story.append(Paragraph(f"Descuento: -${sale.discount:.2f}", right))
    story.append(Paragraph(f"IVA (16%): ${sale.tax:.2f}", right))
    story.append(Paragraph(f"TOTAL: ${sale.total:.2f}", bold_right))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"Pago: {sale.payment_method}", center))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("¡Gracias por su compra!", center))

    doc.build(story)
    return buffer.getvalue()


def generate_sales_excel(sales: list, store_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    header_fill = PatternFill("solid", fgColor="F43F5E")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Reporte de Ventas — {store_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Folio", "Fecha", "Vendedor", "# Productos", "Subtotal", "Descuento", "IVA", "Total", "Método de Pago", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, sale in enumerate(sales, 4):
        ws.cell(row=row, column=1, value=sale.folio)
        ws.cell(row=row, column=2, value=sale.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=3, value=sale.seller.full_name if sale.seller else "")
        ws.cell(row=row, column=4, value=len(sale.items))
        ws.cell(row=row, column=5, value=sale.subtotal)
        ws.cell(row=row, column=6, value=sale.discount)
        ws.cell(row=row, column=7, value=sale.tax)
        ws.cell(row=row, column=8, value=sale.total)
        ws.cell(row=row, column=9, value=sale.payment_method)
        ws.cell(row=row, column=10, value=sale.status)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
